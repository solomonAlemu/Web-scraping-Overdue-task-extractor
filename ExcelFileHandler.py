import subprocess
import sys
import importlib.util
import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

class ExcelFileHandler:
    def __init__(self, file_path=None):
        """
        Initialize the ExcelFileHandler with an existing file or create a new workbook.
        
        :param file_path: Path to the Excel file. If None, a new workbook is created.
        """
        if file_path:
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path)
        else:
            self.file_path = None
            self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def load_sheet(self, sheet_name):
        """
        Load a specific sheet by name.
        
        :param sheet_name: Name of the sheet to load.
        :return: None
        """
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
    def set_by_cellvalue_byrange(self, new_value, range_string):
        """
        Set all cells within a specified range to a new value.
        
        :param new_value: The value to set in each cell.
        :param range_string: The range of cells to update (e.g., 'A1:C10').
        :return: Number of cells updated.
        """
        range_cells = self.sheet[range_string]
        updated_cells = 0

        for row in range_cells:
            for cell in row:
                cell.value = new_value
                updated_cells += 1

        return updated_cells
    
    def read_data(self, range_string=None):
        """
        Read data from the current sheet.
        
        :param range_string: Excel range to read (e.g., 'A1:C3'). If None, reads the entire sheet.
        :return: Data as a Pandas DataFrame.
        """
        if range_string:
            data = self.sheet[range_string]
            rows = [[cell.value for cell in row] for row in data]
            return pd.DataFrame(rows)
        else:
            data = self.sheet.values
            return pd.DataFrame(data)

    def write_data(self, data, start_cell="A1", sheet_name=None):
        """
        Write data to the sheet, starting from the given cell.
        
        :param data: Data to write (Pandas DataFrame, list of lists, or single value).
        :param start_cell: Start cell in Excel notation (e.g., 'B2').
        :param sheet_name: Sheet to write data to. If None, writes to the active sheet.
        :return: None
        """
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(title=sheet_name)
            self.sheet = self.workbook[sheet_name]

        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
        
        if isinstance(data, pd.DataFrame):
            data = data.values.tolist()

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                self.sheet.cell(row=start_row + i, column=start_col + j, value=value)

    def write_string_data(self, data, start_cell="A1", sheet_name=None):
            """
            Write data to the sheet, starting from the given cell.
            
            :param data: Data to write (Pandas DataFrame, list of lists, single string, or single value).
            :param start_cell: Start cell in Excel notation (e.g., 'B2').
            :param sheet_name: Sheet to write data to. If None, writes to the active sheet.
            :return: None
            """
            if sheet_name:
                if sheet_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(title=sheet_name)
                self.sheet = self.workbook[sheet_name]
    
            start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
            
            if isinstance(data, str):
                # Write the string data to a single cell
                self.sheet.cell(row=start_row, column=start_col, value=data)
            elif isinstance(data, pd.DataFrame):
                data = data.values.tolist()
                for i, row in enumerate(data):
                    for j, value in enumerate(row):
                        self.sheet.cell(row=start_row + i, column=start_col + j, value=value)
            elif isinstance(data, list):
                for i, row in enumerate(data):
                    for j, value in enumerate(row):
                        self.sheet.cell(row=start_row + i, column=start_col + j, value=value)
            else:
                # Handle single value input
                self.sheet.cell(row=start_row, column=start_col, value=data)

    # ... (Other methods remain unchanged)

    def update_cell(self, cell, value):
        """
        Update a specific cell with a new value.
        
        :param cell: Cell to update (e.g., 'A1').
        :param value: New value for the cell.
        :return: None
        """
        self.sheet[cell] = value
        
    def update_by_cellvalue_byrange(self, search_value, new_value, range_string):
        """
        Update cells within a specified range if they match a given value.
        
        :param search_value: The value to search for.
        :param new_value: The value to replace the search_value with.
        :param range_string: The range of cells to search within (e.g., 'A1:C10').
        :return: Number of cells updated.
        """
        range_cells = self.sheet[range_string]
        updated_cells = 0

        for row in range_cells:
            for cell in row:
                if cell.value == search_value:
                    cell.value = new_value
                    updated_cells += 1

        return updated_cells
    
    def format_cell(self, cell, font_name=None, font_size=None, bold=None, color=None, fill_color=None):
        """
        Apply formatting to a specific cell.
        
        :param cell: Cell to format (e.g., 'A1').
        :param font_name: Font name (e.g., 'Arial').
        :param font_size: Font size (e.g., 12).
        :param bold: Boolean indicating whether the text should be bold.
        :param color: Font color in hex (e.g., 'FF0000' for red).
        :param fill_color: Cell background color in hex (e.g., 'FFFF00' for yellow).
        :return: None
        """
        cell_obj = self.sheet[cell]
        if font_name or font_size or bold or color:
            cell_obj.font = Font(name=font_name, size=font_size, bold=bold, color=color)
        if fill_color:
            cell_obj.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    def save(self, file_path=None):
        """
        Save the workbook to a file. If only a directory is provided, a default filename is used.
        
        :param file_path: Path to save the file. If None, saves to the original file.
        :return: None
        """
        if file_path:
            # If the path is a directory, append a default filename
            if os.path.isdir(file_path):
                file_path = os.path.join(file_path, "data.xlsx")
            
            # Ensure the file path is not a directory
            if os.path.isdir(file_path):
                raise ValueError(f"The provided path '{file_path}' is a directory, not a file path.")
            
            self.workbook.save(file_path)
        elif self.file_path:
            self.workbook.save(self.file_path)
        else:
            raise ValueError("No file path specified to save the workbook.")

    def export_sheet(self, sheet_name, file_path):
        """
        Export a specific sheet to a new Excel file.
        
        :param sheet_name: Name of the sheet to export.
        :param file_path: Path to the new Excel file.
        :return: None
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
        
        new_workbook = openpyxl.Workbook()
        source_sheet = self.workbook[sheet_name]
        target_sheet = new_workbook.active
        target_sheet.title = sheet_name
        
        for row in source_sheet.iter_rows(values_only=True):
            target_sheet.append(row)
        
        new_workbook.save(file_path)

    def add_sheet(self, sheet_name):
        """
        Add a new sheet to the workbook.
        
        :param sheet_name: Name of the new sheet.
        :return: None
        """
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} already exists.")
        self.workbook.create_sheet(title=sheet_name)
        
    def rename_sheet(self, old_name, new_name):
        """
        Rename an existing sheet in the workbook.
        
        :param old_name: The current name of the sheet to rename.
        :param new_name: The new name for the sheet.
        :return: None
        """
        if old_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet {old_name} does not exist.")
        if new_name in self.workbook.sheetnames:
            raise ValueError(f"Sheet {new_name} already exists.")
        
        sheet = self.workbook[old_name]
        sheet.title = new_name
    def close(self):
        """
        Close the workbook if open.
        
        :return: None
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None       
 # Example usage
if __name__ == "__main__":     
    
 
    # Initialize the handler with a new workbook
    excel_handler = ExcelFileHandler()
    
    # Add a new sheet named 'DataSheet'
    excel_handler.add_sheet('DataSheet')
    print("Sheet 'DataSheet' added successfully.")
    
    # Load the new sheet
    excel_handler.load_sheet('DataSheet')
    
    # Write some data to the new sheet
    data = [['Name', 'Age', 'City'], ['Alice', 30, 'New York'], ['Bob', 25, 'Los Angeles']]
    excel_handler.write_data(data, start_cell='A1')
    
    # Ask the user for a directory path to save the workbook
    #  user_dir_path = input("Please provide the directory path to save the Excel file (e.g., 'C:/Users/solomonal/Downloads'): ")
    
    # Save the workbook to the user-provided directory with a default filename
    excel_handler.save('C:/Users/solomonal/Downloads')
